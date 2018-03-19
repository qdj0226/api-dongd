#-*- coding:utf-8 -*-
__author__ = "dongd"
__datetime__ = r"2018\3\6 0006"
__software__ = "PyCharm"

from openpyxl import load_workbook
from openpyxl.styles import Font
import time

class ResolveXL():
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.fnot = Font(color=None)
        self.RGBDict = {
            "red": "FFFF3030",
            "green": "FF008B00"
        }

    def getWorkBook(self, filepath):
        '''实例化worknook并返回该实例'''
        try:
            self.workbook = load_workbook(filename=filepath)
        except Exception as e:
            raise e
        self.excelFile = filepath
        return self.workbook

    def getSheetByName(self,sheetname):
        '''通过sheet名称获取sheet实例'''
        try:
            sheet = self.workbook[sheetname]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,index):
        '''通过sheet索引号获取sheet实例，下标0开始'''
        try:
            sheet = self.workbook.worksheets[index-1]
            return sheet
        except Exception as e:
            raise e

    def getRowMax(self,sheet):
        '''获取最大行数'''
        return sheet.max_row

    def getColMax(self,sheet):
        '''获取最大列数'''
        return sheet.max_column

    def getRowMin(self,sheet):
        '''获取最小行数'''
        return sheet.min_row

    def getColMin(self,sheet):
        '''获取最小列数'''
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        '''获取行编码（如”A1“），下标从1开始'''
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception as e:
            raise e

    def getCol(self,sheet,colNo):
        '''获取列编码（如”A1“），下标从1开始'''
        try:
            return list(sheet.columns)[colNo-1]
        except Exception as e:
            raise e


    def getCellValue(self,sheet,rowNo,colNo):
        '''获取单元格值'''
        try:
            value = sheet.cell(row=rowNo,column = colNo)
            return value
        except Exception as e:
            raise e

    def getCellObject(self,sheet,rowNo,colNo):
        '''获取单元格对象'''
        try:
            object = sheet.cell(row=rowNo,column=colNo)
            return object
        except Exception as e:
            raise e

    def writeCell(self,sheet,value,coordinate=None,rowNo=None,colNo=None,style=None):
        '''
        向单元格中写入数据
        :param sheet: sheet实例
        :param value: 写入的值
        :param coordinate: 编码定位，如“A1”
        :param rowNo: 行索引号,下标1开始
        :param colNo: 列索引号，下标1开始
        :param style: 设置单元格color属性
        '''
        if coordinate is not None:
            try:
                sheet[coordinate].value = value
                if style is not None:
                    sheet[coordinate].font = Font(color=self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = value
                if style is not None:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("请输入正确的索引信息")

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colNo=None):
        '''向单元格中写入当前日期时间'''
        date = time.time()
        date = time.localtime(date)
        date = time.strftime("%Y-%m-%d %H:%M:%S", date)
        if coordinate is not None:
            try:
                sheet[coordinate] = date
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo,column=colNo).value = date
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("请输入正确的索引信息")

if __name__ == '__main__':
    from config.configFile import dataexcelfile
    rxl = ResolveXL()
    wb = rxl.getWorkBook(dataexcelfile)
    ws = rxl.getSheetByIndex(1)
    print(rxl.getRow(ws,1))
    print(rxl.getColMax(ws))
    print(rxl.getRowMax(ws))
    print(rxl.getColMin(ws))
    print(rxl.getRowMin(ws))
    print(rxl.getCol(ws,1))
    print(ws.cell(1,1).value)
